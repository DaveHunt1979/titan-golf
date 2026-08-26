#!/usr/bin/env python3
"""
Build ONE clean, deduplicated golf-course master workbook from:
  - TITAN_GLOBAL_GOLF_COURSE_MASTER_AUDITED.xlsx (England/Scotland/Ireland&NI/
    Portugal/Turkey/Orlando — 755 "Courses" rows)
  - Wales / France / Spain source workbooks (not yet folded into the master)

All four are read-only inputs. Nothing is ever modified in place — this
writes a NEW workbook (TITAN_GLOBAL_GOLF_COURSE_MASTER_UNIFIED.xlsx).

IMPORTANT FINDING (discovered while building this): the master's OWN 755
"Courses" rows are not all distinct physical courses. At least 118 of them
are "clean" rows (e.g. Course ID "MID_KENT", 4 tees already attached) sitting
alongside redundant per-tee "shadow" rows (e.g. "MID_KENT__YELLOW_M") whose
Official Course Name has the tee/gender baked in via an em dash
("The Mid Kent Golf Club — Yellow M Rated Configuration") and whose tee/hole
data is a byte-identical duplicate of a tee the clean row already has. A few
courses (e.g. West Cliffs) have NO clean aggregate row at all — only 7
separate per-tee shadow rows ("West Cliffs — Tee 1".."Tee 7") that all need
merging into one course with 7 distinct tees.

This is exactly the same shape of problem the Wales/France/Spain merge was
built to solve, so this script applies ONE grouping/dedup pass across ALL
four sources combined, rather than trusting the master's 6 regions as
pre-clean. Ambiguous cases (conflicting hole counts, conflicting region,
incomplete hole data) are excluded from the output and written to
needs_review.csv for a human to resolve — nothing is silently guessed.

Usage:
    python3 scripts/merge_course_master.py
"""
import csv
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
GC = ROOT / "screenshots" / "Golf courses"
MASTER_PATH = GC / "UK & Ireland" / "TITAN_GLOBAL_GOLF_COURSE_MASTER_AUDITED.xlsx"
OUT_PATH = GC / "UK & Ireland" / "TITAN_GLOBAL_GOLF_COURSE_MASTER_UNIFIED.xlsx"
REVIEW_PATH = ROOT / "scripts" / "needs_review.csv"

MASTER_COURSES_HDR = [
    "Country / Region", "Course ID", "Official Course Name", "Club / Venue",
    "Region / County", "Town / Area", "Course / Layout", "Holes", "Status",
    "Primary Source", "Last Verified", "Notes / Source Quality",
    "Distance Unit", "Source Workbook",
]
MASTER_TEES_HDR = [
    "Country / Region", "Course ID", "Official Course Name", "Tee Name",
    "Gender", "Par", "Total Distance", "Distance Unit", "Course Rating",
    "Slope Rating", "Source", "Rating Status",
]
MASTER_HOLES_HDR = [
    "Country / Region", "Course ID", "Official Course Name", "Tee Name",
    "Gender", "Hole", "Distance", "Distance Unit", "Par", "Stroke Index",
    "Source",
]

EM_DASH = "—"

# Each entry describes one input source: where to read it, how its columns
# map onto the canonical internal field names, and whether country/distance
# unit are fixed per file or read per-row (master mixes 6 countries).
SOURCES = [
    {
        "label": "master",
        "file": MASTER_PATH,
        "sheet_names": {"courses": "Courses", "tees": "Tee Ratings", "holes": "Hole Data"},
        "country_col": "Country / Region",     # per-row
        "distance_unit_col": "Distance Unit",  # per-row
        "courses_map": {
            "Country / Region": "country", "Course ID": "course_id",
            "Official Course Name": "name", "Club / Venue": "club",
            "Region / County": "region", "Town / Area": "town",
            "Course / Layout": "layout", "Holes": "holes", "Status": "status",
            "Primary Source": "source", "Last Verified": "verified",
            "Notes / Source Quality": "notes", "Distance Unit": "distance_unit",
        },
        "tees_map": {
            "Course ID": "course_id", "Tee Name": "tee_name", "Gender": "gender",
            "Par": "par", "Total Distance": "distance", "Course Rating": "course_rating",
            "Slope Rating": "slope_rating", "Source": "source",
        },
        "holes_map": {
            "Course ID": "course_id", "Tee Name": "tee_name", "Gender": "gender",
            "Hole": "hole", "Distance": "distance", "Par": "par",
            "Stroke Index": "stroke_index", "Source": "source",
        },
    },
    {
        "label": "Wales",
        "file": GC / "UK & Ireland" / "Wales_MASTER_COMPLETE_ONLY_GOLFBREAKS_YGT_COMPLETE_20.xlsx",
        "sheet_names": {"courses": "Courses", "tees": "Tee Ratings", "holes": "Hole Data"},
        "country_const": "Wales", "distance_unit_const": "yd",
        "courses_map": {
            "Course ID": "course_id", "Official Course Name": "name",
            "Club / Venue": "club", "County / Region": "region", "Town": "town",
            "Layout": "layout", "Holes": "holes", "Status": "status",
            "Source": "source", "Verified": "verified", "Notes": "notes",
        },
        "tees_map": {
            "Course ID": "course_id", "Tee": "tee_name", "Gender": "gender", "Par": "par",
            "Distance": "distance", "Course Rating": "course_rating",
            "Slope Rating": "slope_rating", "Source": "source",
        },
        "holes_map": {
            "Course ID": "course_id", "Tee": "tee_name", "Gender": "gender", "Hole": "hole",
            "Distance": "distance", "Par": "par", "Stroke Index": "stroke_index", "Source": "source",
        },
    },
    {
        "label": "France",
        "file": GC / "euro courses" / "France_GOLFBREAKS_YGT_FRANCE_FINAL_AUDIT_PROGRESS_59.xlsx",
        "sheet_names": {"courses": "Courses", "tees": "Tee Ratings", "holes": "Hole Data"},
        "country_const": "France", "distance_unit_const": "yd",
        "courses_map": {
            "Course ID": "course_id", "Official Course Name": "name",
            "Club Name": "club", "Region": "region", "Town/Area": "town",
            "Course / Layout": "layout", "Holes": "holes",
            "Verification Status": "status", "Primary Source": "source",
            "Last Verified": "verified", "Source Quality": "notes",
        },
        "tees_map": {
            "Course ID": "course_id", "Tee Name": "tee_name", "Gender": "gender",
            "Total Par": "par", "Total Length (yds)": "distance",
            "Course Rating": "course_rating", "Slope Rating": "slope_rating", "Source": "source",
        },
        "holes_map": {
            "Course ID": "course_id", "Tee Name": "tee_name", "Gender": "gender", "Hole": "hole",
            "Yardage": "distance", "Par": "par", "Stroke Index": "stroke_index", "Source": "source",
        },
    },
    {
        "label": "Spain",
        "file": GC / "euro courses" / "Spain_GOLFBREAKS_YGT_FINAL_PUSH_97(1).xlsx",
        "sheet_names": {"courses": "Courses", "tees": "Tee Ratings", "holes": "Hole Data"},
        "country_const": "Spain", "distance_unit_const": "yd",
        "courses_map": {
            "Course ID": "course_id", "Official Course Name": "name",
            "Club Name": "club", "Region": "region", "Town/Area": "town",
            "Course / Layout": "layout", "Holes": "holes",
            "Verification Status": "status", "Primary Source": "source",
            "Last Verified": "verified", "Source Quality": "notes",
        },
        "tees_map": {
            "Course ID": "course_id", "Tee Name": "tee_name", "Gender": "gender",
            "Total Par": "par", "Total Length (yds)": "distance",
            "Course Rating": "course_rating", "Slope Rating": "slope_rating", "Source": "source",
        },
        "holes_map": {
            "Course ID": "course_id", "Tee Name": "tee_name", "Gender": "gender", "Hole": "hole",
            "Yardage": "distance", "Par": "par", "Stroke Index": "stroke_index", "Source": "source",
        },
    },
    {
        # Rick-supplied, dropped 2026-08-25 directly in screenshots/ (not yet
        # filed under a country subfolder). Same column shape as France/Spain
        # above; kept as an ADDITIONAL source rather than replacing the older
        # England rows already inside "master" -- the grouping/dedup pass
        # picks whichever member has more tees as canonical and merges every
        # valid tee across all duplicate sources, so adding this alongside
        # master's existing England rows is safe by construction.
        "label": "England (Rick, 2026-08-25)",
        "file": ROOT / "screenshots" / "TITAN_ENGLAND_API_SAFE_MASTER_PROGRESS_33_CONTINUED_PUSH.xlsx",
        "sheet_names": {"courses": "Courses", "tees": "Tee Ratings", "holes": "Hole Data"},
        "country_const": "England", "distance_unit_const": "yd",
        "courses_map": {
            "Course ID": "course_id", "Official Course Name": "name",
            "Club Name": "club", "County": "region", "Town/Area": "town",
            "Course / Layout": "layout", "Holes": "holes",
            "Verification Status": "status", "Primary Source": "source",
            "Last Verified": "verified", "Source Quality": "notes",
        },
        "tees_map": {
            "Course ID": "course_id", "Tee Name": "tee_name", "Gender": "gender",
            "Total Par": "par", "Total Length (yds)": "distance",
            "Course Rating": "course_rating", "Slope Rating": "slope_rating", "Source": "source",
        },
        "holes_map": {
            "Course ID": "course_id", "Tee Name": "tee_name", "Gender": "gender", "Hole": "hole",
            "Yardage": "distance", "Par": "par", "Stroke Index": "stroke_index", "Source": "source",
        },
    },
    {
        # Rick-supplied, dropped 2026-08-25. Additional Spain source, same
        # reasoning as England above -- kept alongside the older Spain file
        # rather than replacing it.
        "label": "Spain (Rick, 2026-08-25)",
        "file": ROOT / "screenshots" / "Spain_GOLFBREAKS_YGT_TEE_SET_AUDIT_FINAL.xlsx",
        "sheet_names": {"courses": "Courses", "tees": "Tee Ratings", "holes": "Hole Data"},
        "country_const": "Spain", "distance_unit_const": "yd",
        "courses_map": {
            "Course ID": "course_id", "Official Course Name": "name",
            "Club Name": "club", "Region": "region", "Town/Area": "town",
            "Course / Layout": "layout", "Holes": "holes",
            "Verification Status": "status", "Primary Source": "source",
            "Last Verified": "verified", "Source Quality": "notes",
        },
        "tees_map": {
            "Course ID": "course_id", "Tee Name": "tee_name", "Gender": "gender",
            "Total Par": "par", "Total Length (yds)": "distance",
            "Course Rating": "course_rating", "Slope Rating": "slope_rating", "Source": "source",
        },
        "holes_map": {
            "Course ID": "course_id", "Tee Name": "tee_name", "Gender": "gender", "Hole": "hole",
            "Yardage": "distance", "Par": "par", "Stroke Index": "stroke_index", "Source": "source",
        },
    },
    {
        # Rick-supplied, dropped 2026-08-25. Turkey was never wired into this
        # merge at all before now (neither this file nor the older
        # euro-courses Turkey workbook) -- adding both; the older one may
        # still hold courses this newer "TEE SET AUDIT FINISHED" pass didn't
        # re-cover. Distances are explicitly labelled Distance_m (unambiguous
        # metres), unlike the French/Spanish sources which store raw numbers
        # under a nominal "yd" label -- kept accurate here since the source
        # itself removes the ambiguity.
        "label": "Turkey (Rick, 2026-08-25)",
        "file": ROOT / "screenshots" / "Turkey_MASTER_Golf_Course_Database_TEE_SET_AUDIT_FINISHED.xlsx",
        "sheet_names": {"courses": "Courses", "tees": "Tee Ratings", "holes": "Hole Data"},
        "country_const": "Turkey", "distance_unit_const": "m",
        "courses_map": {
            "Course_ID": "course_id", "Course_Name": "name", "Club": "club",
            "Region": "region", "Town": "town", "Layout": "layout", "Holes": "holes",
            "Status": "status", "Source_URL": "source", "Verified_Date": "verified",
            "Notes": "notes",
        },
        "tees_map": {
            "Course_ID": "course_id", "Tee": "tee_name", "Gender": "gender", "Par": "par",
            "Distance_m": "distance", "Course_Rating": "course_rating",
            "Slope_Rating": "slope_rating", "Source_URL": "source",
        },
        "holes_map": {
            "Course_ID": "course_id", "Tee": "tee_name", "Gender": "gender", "Hole": "hole",
            "Distance_m": "distance", "Par": "par", "Stroke_Index": "stroke_index",
            "Source_URL": "source",
        },
    },
    {
        "label": "Turkey (old, previously never wired in)",
        "file": GC / "euro courses" / "Turkey_MASTER_Golf_Course_Database_WITH_DALAMAN_EXCEPTION_18.xlsx",
        "sheet_names": {"courses": "Courses", "tees": "Tee Ratings", "holes": "Hole Data"},
        "country_const": "Turkey", "distance_unit_const": "m",
        "courses_map": {
            "Course_ID": "course_id", "Course_Name": "name", "Club": "club",
            "Region": "region", "Town": "town", "Layout": "layout", "Holes": "holes",
            "Status": "status", "Source_URL": "source", "Verified_Date": "verified",
            "Notes": "notes",
        },
        "tees_map": {
            "Course_ID": "course_id", "Tee": "tee_name", "Gender": "gender", "Par": "par",
            "Distance_m": "distance", "Course_Rating": "course_rating",
            "Slope_Rating": "slope_rating", "Source_URL": "source",
        },
        "holes_map": {
            "Course_ID": "course_id", "Tee": "tee_name", "Gender": "gender", "Hole": "hole",
            "Distance_m": "distance", "Par": "par", "Stroke_Index": "stroke_index",
            "Source_URL": "source",
        },
    },
]


def strip_tee_suffix(name: str) -> str:
    """Cut everything from the first EM DASH onward. Every source uses a
    plain ASCII hyphen for genuine distinct-layout naming (e.g. 'Woburn Golf
    Club - The Duchess', 'Gleneagles - PGA Centenary') and reserves the em
    dash specifically for a baked-in tee/gender/'Rated Configuration'
    suffix (e.g. 'West Cliffs — Tee 4', 'The Mid Kent Golf Club — Yellow M
    Rated Configuration') — confirmed by direct inspection of both the
    master and the new source files, not assumed."""
    if EM_DASH in name:
        return name.split(EM_DASH, 1)[0].strip()
    return name.strip()


def normalize_key(name: str) -> str:
    s = name.lower()
    s = s.replace("&", "and")
    s = re.sub(r"[’']", "", s)
    s = re.sub(r"[^a-z0-9]+", " ", s)
    s = re.sub(r"\bthe\b", " ", s)  # "The X - Heritage" and "X - The Heritage" must key the same
    return re.sub(r"\s+", " ", s).strip()


def slugify(name: str) -> str:
    s = name.upper()
    s = s.replace("&", "AND")
    s = re.sub(r"[’']", "", s)
    s = re.sub(r"[^A-Z0-9]+", "_", s)
    return re.sub(r"_+", "_", s).strip("_")


def load_rows(path, sheet, colmap):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(header)}
    out = []
    for raw in rows[1:]:
        if raw is None or all(v is None for v in raw):
            continue
        rec = {}
        for src_col, field in colmap.items():
            i = idx.get(src_col)
            rec[field] = raw[i] if i is not None and i < len(raw) else None
        if rec.get("course_id") is None:
            continue
        out.append(rec)
    return out


def load_source(cfg):
    """Returns (courses, tees_by_course_id, holes_by_course_id) with every
    course record carrying a resolved 'country' and 'distance_unit', and a
    'source_workbook' tag for provenance."""
    courses = load_rows(cfg["file"], cfg["sheet_names"]["courses"], cfg["courses_map"])
    tees = load_rows(cfg["file"], cfg["sheet_names"]["tees"], cfg["tees_map"])
    holes = load_rows(cfg["file"], cfg["sheet_names"]["holes"], cfg["holes_map"])

    for c in courses:
        c["country"] = c.get("country") or cfg.get("country_const")
        c["distance_unit"] = c.get("distance_unit") or cfg.get("distance_unit_const") or "yd"
        c["source_workbook"] = cfg["file"].name
        c["_base_name"] = strip_tee_suffix(c.get("name") or "")

    tees_by_id = defaultdict(list)
    for t in tees:
        tees_by_id[t["course_id"]].append(t)
    holes_by_id = defaultdict(list)
    for h in holes:
        holes_by_id[h["course_id"]].append(h)

    return courses, tees_by_id, holes_by_id


def main():
    all_source_courses = []
    tees_by_id = {}
    holes_by_id = {}
    review_rows = []

    for cfg in SOURCES:
        courses, t_by_id, h_by_id = load_source(cfg)
        all_source_courses.extend(courses)
        tees_by_id.update(t_by_id)
        holes_by_id.update(h_by_id)
        print(f"loaded {cfg['label']}: {len(courses)} raw course rows, "
              f"{sum(len(v) for v in t_by_id.values())} tee rows, "
              f"{sum(len(v) for v in h_by_id.values())} hole rows")

    # Group ALL sources' raw Courses rows together by (country, base name) —
    # this is what collapses master's own redundant shadow rows (MID_KENT +
    # MID_KENT__YELLOW_M etc.) exactly the same way it collapses the
    # tee-suffixed rows from Wales/France/Spain.
    groups = defaultdict(list)
    for c in all_source_courses:
        key = (c["country"], normalize_key(c["_base_name"]))
        groups[key].append(c)

    out_courses, out_tees, out_holes = [], [], []
    used_ids = set()

    for (country, _key), members in sorted(groups.items(), key=lambda kv: (kv[0][0] or "", kv[1][0]["_base_name"])):
        base_name = members[0]["_base_name"]
        holes_counts = {m.get("holes") for m in members if m.get("holes") is not None}
        regions = {(m.get("region") or "").strip().lower() for m in members if m.get("region")}

        if len(holes_counts) > 1:
            review_rows.append({
                "region": country, "key": base_name,
                "reason": f"conflicting Holes count across grouped rows: {holes_counts}",
                "course_ids": ",".join(str(m["course_id"]) for m in members),
            })
            continue
        if len(regions) > 1:
            review_rows.append({
                "region": country, "key": base_name,
                "reason": f"same name, conflicting Region/Town across grouped rows: {regions}",
                "course_ids": ",".join(str(m["course_id"]) for m in members),
            })
            continue

        expected_holes = next(iter(holes_counts), None)

        # Prefer, as the metadata donor, whichever member already carries the
        # most tees under its own course_id (the "clean" row, if one
        # exists) — falls back to longest notes text as a tiebreak.
        def tee_count(m):
            return len(tees_by_id.get(m["course_id"], []))
        canonical = max(members, key=lambda m: (tee_count(m), len(str(m.get("notes") or ""))))
        base_name = canonical["_base_name"]  # deterministic: the metadata donor's own name

        unique_ids = list(dict.fromkeys(m["course_id"] for m in members))

        candidate_tees = []
        seen_tees = set()
        for cid in unique_ids:
            for t in tees_by_id.get(cid, []):
                dedupe_key = (t.get("tee_name"), t.get("gender"), t.get("par"),
                              t.get("distance"), t.get("course_rating"), t.get("slope_rating"))
                if dedupe_key in seen_tees:
                    continue
                seen_tees.add(dedupe_key)
                candidate_tees.append(t)

        holes_by_tee = defaultdict(list)
        seen_holes = set()
        for cid in unique_ids:
            for h in holes_by_id.get(cid, []):
                dedupe_key = (h.get("tee_name"), h.get("gender"), h.get("hole"),
                              h.get("distance"), h.get("par"), h.get("stroke_index"))
                if dedupe_key in seen_holes:
                    continue
                seen_holes.add(dedupe_key)
                holes_by_tee[(h.get("tee_name"), h.get("gender"))].append(h)

        valid_tees = []
        for t in candidate_tees:
            tk = (t.get("tee_name"), t.get("gender"))
            tee_holes = holes_by_tee.get(tk, [])
            n = len(tee_holes)
            if expected_holes and n != expected_holes:
                review_rows.append({
                    "region": country, "key": base_name,
                    "reason": f"incomplete hole data for tee {tk}: {n}/{expected_holes} holes — excluded",
                    "course_ids": ",".join(str(m['course_id']) for m in members),
                })
                continue
            bad = [h for h in tee_holes
                   if not (isinstance(h.get("par"), int) and 3 <= h["par"] <= 6)
                   or not (isinstance(h.get("stroke_index"), int) and 1 <= h["stroke_index"] <= 18)
                   or not (isinstance(h.get("hole"), int) and 1 <= h["hole"] <= 18)]
            if bad:
                review_rows.append({
                    "region": country, "key": base_name,
                    "reason": f"invalid par/stroke_index/hole value for tee {tk}: {bad} — excluded",
                    "course_ids": ",".join(str(m['course_id']) for m in members),
                })
                continue
            # A genuine 18-hole layout essentially never repeats its exact
            # (distance, par) sequence between front and back nine — this
            # signature means a scraper duplicated one 9 onto the other
            # (found via the Dorset Golf and Country Club "Yellow" tee,
            # where holes 1-9 are byte-identical to holes 10-18).
            #
            # EXCEPTION: some real venues genuinely are a 9-hole course
            # played twice for an "18-hole" round, which produces this exact
            # signature legitimately (real-world-confirmed 2026-08-26: Crane
            # Valley, Leeds Castle, Lullingstone Park - Valley Course, Meon
            # Valley Marriott, Parley Golf Centre, Westridge Golf Centre,
            # Pine Cliffs). Course names below are exempted from this check;
            # everything else still gets flagged and excluded by default.
            NINE_HOLE_PLAYED_TWICE_VENUES = {
                "crane valley golf club", "leeds castle golf course",
                "lullingstone park golf course valley course",
                "meon valley marriott hotel and country club",
                "parley golf centre",
                "westridge golf centre", "pine cliffs golf course",
            }
            by_hole = {h["hole"]: (h.get("distance"), h.get("par")) for h in tee_holes}
            front9 = [by_hole.get(i) for i in range(1, 10)]
            back9 = [by_hole.get(i) for i in range(10, 19)]
            if (len(by_hole) == 18 and front9 == back9
                    and normalize_key(base_name) not in NINE_HOLE_PLAYED_TWICE_VENUES):
                review_rows.append({
                    "region": country, "key": base_name,
                    "reason": f"front/back nine (distance, par) sequence identical for tee {tk} — "
                              f"looks like duplicated scrape data, not a real 18-hole layout — excluded",
                    "course_ids": ",".join(str(m['course_id']) for m in members),
                })
                continue
            valid_tees.append(t)

        if not valid_tees:
            review_rows.append({
                "region": country, "key": base_name,
                "reason": "no complete tee configuration survived — course excluded entirely",
                "course_ids": ",".join(str(m["course_id"]) for m in members),
            })
            continue

        new_id = slugify(base_name) or f"COURSE_{len(used_ids)+1}"
        n = 2
        base_new_id = new_id
        while new_id in used_ids:
            new_id = f"{base_new_id}_{n}"
            n += 1
        used_ids.add(new_id)

        out_courses.append({
            "Country / Region": country,
            "Course ID": new_id,
            "Official Course Name": base_name,
            "Club / Venue": canonical.get("club"),
            "Region / County": canonical.get("region"),
            "Town / Area": canonical.get("town"),
            "Course / Layout": canonical.get("layout"),
            "Holes": canonical.get("holes"),
            "Status": canonical.get("status"),
            "Primary Source": canonical.get("source"),
            "Last Verified": canonical.get("verified"),
            "Notes / Source Quality": canonical.get("notes"),
            "Distance Unit": canonical.get("distance_unit"),
            "Source Workbook": canonical.get("source_workbook"),
        })

        for t in valid_tees:
            out_tees.append({
                "Country / Region": country,
                "Course ID": new_id,
                "Official Course Name": base_name,
                "Tee Name": t.get("tee_name"),
                "Gender": t.get("gender"),
                "Par": t.get("par"),
                "Total Distance": t.get("distance"),
                "Distance Unit": canonical.get("distance_unit"),
                "Course Rating": t.get("course_rating"),
                "Slope Rating": t.get("slope_rating"),
                "Source": t.get("source"),
                "Rating Status": "VERIFIED" if len(members) == 1 else "MERGED",
            })
            for h in holes_by_tee.get((t.get("tee_name"), t.get("gender")), []):
                out_holes.append({
                    "Country / Region": country,
                    "Course ID": new_id,
                    "Official Course Name": base_name,
                    "Tee Name": h.get("tee_name"),
                    "Gender": h.get("gender"),
                    "Hole": h.get("hole"),
                    "Distance": h.get("distance"),
                    "Distance Unit": canonical.get("distance_unit"),
                    "Par": h.get("par"),
                    "Stroke Index": h.get("stroke_index"),
                    "Source": h.get("source"),
                })

    out_courses.sort(key=lambda c: (c["Country / Region"] or "", c["Official Course Name"] or ""))
    out_tees.sort(key=lambda t: (t["Country / Region"] or "", t["Course ID"] or "", t["Tee Name"] or ""))
    out_holes.sort(key=lambda h: (h["Country / Region"] or "", h["Course ID"] or "",
                                   h["Tee Name"] or "", h["Hole"] or 0))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def write_sheet(name, hdr, rows):
        ws = wb.create_sheet(name)
        ws.append(hdr)
        for r in rows:
            ws.append([r.get(h) for h in hdr])

    write_sheet("Courses", MASTER_COURSES_HDR, out_courses)
    write_sheet("Tee Ratings", MASTER_TEES_HDR, out_tees)
    write_sheet("Hole Data", MASTER_HOLES_HDR, out_holes)

    audit_ws = wb.create_sheet("Audit Summary")
    audit_ws.append(["Country / Region", "Deduplicated Course Records",
                      "Qualified Tee Configurations", "Complete Hole Rows", "Source Workbook(s)"])
    by_country = defaultdict(lambda: [0, 0, 0])
    for c in out_courses:
        by_country[c["Country / Region"]][0] += 1
    for t in out_tees:
        by_country[t["Country / Region"]][1] += 1
    for h in out_holes:
        by_country[h["Country / Region"]][2] += 1
    for country in sorted(by_country):
        n_c, n_t, n_h = by_country[country]
        audit_ws.append([country, n_c, n_t, n_h, ""])
    audit_ws.append([])
    audit_ws.append(["MASTER TOTALS", "", "", "", ""])
    audit_ws.append(["Course records", len(out_courses), "", "", ""])
    audit_ws.append(["Tee configurations", len(out_tees), "", "", ""])
    audit_ws.append(["Hole rows", len(out_holes), "", "", ""])
    audit_ws.append(["Regions merged", len(by_country), "", "", ""])
    audit_ws.append(["Raw source Courses rows processed", len(all_source_courses), "", "", ""])
    audit_ws.append(["Items flagged for review (excluded)", len(review_rows), "", "", ""])

    wb.save(OUT_PATH)
    print(f"\nWrote unified workbook: {OUT_PATH}")
    print(f"Totals: {len(out_courses)} courses, {len(out_tees)} tees, {len(out_holes)} hole rows "
          f"(from {len(all_source_courses)} raw source rows)")

    with open(REVIEW_PATH, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["region", "key", "reason", "course_ids"])
        w.writeheader()
        w.writerows(review_rows)
    print(f"Flagged {len(review_rows)} item(s) for review -> {REVIEW_PATH}")


if __name__ == "__main__":
    main()
