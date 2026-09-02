import openpyxl, glob, json, re, sys
from collections import defaultdict

FOLDER = "/Users/davehunt/Desktop/titan-golf/screenshots/Golf courses"

def norm_key(h):
    if h is None: return ''
    h = str(h).lower().strip()
    h = h.replace('_', ' ').replace('/', ' ')
    h = re.sub(r'[^a-z0-9 ]', '', h)
    h = re.sub(r'\s+', ' ', h).strip()
    return h

ALIASES = {
    'course id': 'course_id',
    'course_id': 'course_id',
    'tee config id': 'tee_config_id',
    'official course name': 'course_name',
    'course name': 'course_name',
    'club venue': 'club',
    'club name': 'club',
    'club': 'club',
    'region': 'region',
    'county region': 'region',
    'county': 'region',
    'emirate': 'region',
    'town area': 'town',
    'town': 'town',
    'layout': 'layout',
    'course layout': 'layout',
    'holes': 'holes',
    'tee': 'tee_name',
    'tee name': 'tee_name',
    'gender': 'gender',
    'par': 'par',
    'total par': 'par',
    'distance yds': 'distance',
    'distance': 'distance',
    'distance yd': 'distance',
    'distance m': 'distance_m',
    'total length yds': 'distance',
    'yardage': 'distance',
    'course rating': 'course_rating',
    'slope rating': 'slope_rating',
    'source': 'source',
    'source url': 'source',
    'hole': 'hole',
    'stroke index': 'stroke_index',
}

def hmap(header):
    m = {}
    for i, h in enumerate(header):
        k = norm_key(h)
        k = ALIASES.get(k, None)
        if k and k not in m:
            m[k] = i
    return m

def sheet_rows(ws):
    it = ws.iter_rows(values_only=True)
    header = next(it)
    hm = hmap(header)
    for row in it:
        if row is None or all(v is None for v in row):
            continue
        yield hm, row

def get(hm, row, key, default=None):
    idx = hm.get(key)
    if idx is None or idx >= len(row):
        return default
    v = row[idx]
    return v if v is not None else default

courses_out = {}          # course_name -> {region}
tees_out = {}              # (course_name, tee_name, gender) -> dict  (dedupe within batch)
tee_holes_out = {}         # (course_name, tee_name, gender, hole) -> dict
course_holes_by_course = defaultdict(dict)  # course_name -> tee_name -> {hole: (par, si, dist)}
errors = []
name_collisions = defaultdict(list)  # course_name -> [source files]

files = sorted(glob.glob(f"{FOLDER}/*.xlsx"))
print(f"Found {len(files)} files")

for fpath in files:
    fname = fpath.split('/')[-1]
    is_metres = 'Turkiye' in fname or 'Turkey' in fname
    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)

    # ---- Courses sheet ----
    course_region = {}
    course_id_to_name = {}
    course_holes_count = {}
    if 'Courses' in wb.sheetnames:
        for hm, row in sheet_rows(wb['Courses']):
            cname = get(hm, row, 'course_name')
            cid = get(hm, row, 'course_id')
            if not cname:
                continue
            cname = str(cname).strip()
            if cid is not None:
                course_id_to_name[str(cid).strip()] = cname
            region = get(hm, row, 'region') or get(hm, row, 'town')
            course_region[cname] = region
            holes_n = get(hm, row, 'holes')
            course_holes_count[cname] = holes_n
            if cname in courses_out and cname not in name_collisions:
                pass
            name_collisions[cname].append(fname)
            courses_out.setdefault(cname, {'region': region})

    # ---- Tee Ratings sheet ----
    tee_config_to_key = {}  # England: tee_config_id -> (course_name, tee_name, gender)
    if 'Tee Ratings' in wb.sheetnames:
        for hm, row in sheet_rows(wb['Tee Ratings']):
            cname = get(hm, row, 'course_name')
            cid = get(hm, row, 'course_id')
            if not cname and cid is not None:
                cname = course_id_to_name.get(str(cid).strip())
            if not cname:
                errors.append(f"{fname}/Tee Ratings: row with no resolvable course_name (course_id={cid})")
                continue
            cname = str(cname).strip()
            tee_name = get(hm, row, 'tee_name')
            gender = get(hm, row, 'gender') or ''
            gender = gender.strip().upper() if isinstance(gender, str) else ''
            if gender not in ('M', 'F'):
                gender = ''
            if not tee_name:
                errors.append(f"{fname}/Tee Ratings: missing tee_name for {cname}")
                continue
            tee_name = str(tee_name).strip()
            dist = get(hm, row, 'distance')
            if dist is None and is_metres:
                dist = get(hm, row, 'distance_m')
            if dist is not None and is_metres:
                dist = round(float(dist) * 1.09361)
            par = get(hm, row, 'par')
            cr = get(hm, row, 'course_rating')
            sr = get(hm, row, 'slope_rating')
            source = get(hm, row, 'source')
            key = (cname, tee_name, gender)
            tees_out[key] = {
                'course_name': cname, 'tee_name': tee_name, 'gender': gender,
                'par': int(par) if par is not None else None,
                'total_distance': int(dist) if dist is not None else None,
                'distance_unit': 'yd',
                'course_rating': float(cr) if cr is not None else None,
                'slope_rating': int(sr) if sr is not None else None,
                'source': str(source) if source is not None else None,
                'rating_status': None,
                'source_course_id': str(cid) if cid is not None else None,
            }
            tcid = get(hm, row, 'tee_config_id')
            if tcid is not None:
                tee_config_to_key[str(tcid).strip()] = key

    # ---- Hole Data sheet ----
    if 'Hole Data' in wb.sheetnames:
        for hm, row in sheet_rows(wb['Hole Data']):
            tcid = get(hm, row, 'tee_config_id')
            cname = get(hm, row, 'course_name')
            cid = get(hm, row, 'course_id')
            tee_name = get(hm, row, 'tee_name')
            gender = get(hm, row, 'gender') or ''
            gender = gender.strip().upper() if isinstance(gender, str) else ''
            if gender not in ('M', 'F'):
                gender = ''

            if tcid is not None and (cname is None or tee_name is None):
                # England-style: resolve via Tee Ratings' tee_config_id map
                key = tee_config_to_key.get(str(tcid).strip())
                if key is None:
                    errors.append(f"{fname}/Hole Data: unresolved tee_config_id={tcid}")
                    continue
                cname, tee_name, gender = key
            else:
                if not cname and cid is not None:
                    cname = course_id_to_name.get(str(cid).strip())
                if not cname or not tee_name:
                    errors.append(f"{fname}/Hole Data: unresolved row (course_id={cid}, tee={tee_name})")
                    continue
                cname = str(cname).strip()
                tee_name = str(tee_name).strip()

            hole = get(hm, row, 'hole')
            par = get(hm, row, 'par')
            si = get(hm, row, 'stroke_index')
            dist = get(hm, row, 'distance')
            if dist is None and is_metres:
                dist = get(hm, row, 'distance_m')
            if dist is not None and is_metres:
                dist = round(float(dist) * 1.09361)

            if hole is None or par is None or si is None:
                errors.append(f"{fname}/Hole Data: incomplete row for {cname}/{tee_name} hole={hole}")
                continue
            hole = int(hole)
            par = int(par)
            si = int(si)
            key = (cname, tee_name, gender, hole)
            tee_holes_out[key] = {
                'course_name': cname, 'tee_name': tee_name, 'gender': gender,
                'hole_number': hole, 'distance': int(dist) if dist is not None else None,
                'par': par, 'stroke_index': si,
                'source_course_id': str(cid) if cid is not None else None,
            }
            course_holes_by_course[cname].setdefault(tee_name, {})[hole] = (par, si, dist)
            courses_out.setdefault(cname, {'region': course_region.get(cname)})

    wb.close()
    print(f"parsed {fname}: courses so far={len(courses_out)}, tees so far={len(tees_out)}, tee_holes so far={len(tee_holes_out)}")

print("\n--- totals ---")
print("courses:", len(courses_out))
print("course_tees:", len(tees_out))
print("course_tee_holes:", len(tee_holes_out))
print("parse errors:", len(errors))

with open('scripts/course_parse_errors.txt', 'w') as f:
    f.write('\n'.join(errors))

dupes = {k: v for k, v in name_collisions.items() if len(set(v)) > 1}
print("course names appearing in >1 source file:", len(dupes))
with open('scripts/course_name_collisions.txt', 'w') as f:
    for k, v in dupes.items():
        f.write(f"{k}: {sorted(set(v))}\n")

# Save intermediate parsed data
with open('scripts/parsed_courses.json', 'w') as f:
    json.dump(courses_out, f)
with open('scripts/parsed_tees.json', 'w') as f:
    json.dump(list(tees_out.values()), f)
with open('scripts/parsed_tee_holes.json', 'w') as f:
    json.dump(list(tee_holes_out.values()), f)
with open('scripts/parsed_course_holes_raw.json', 'w') as f:
    # course -> tee -> hole -> [par, si, dist]
    serializable = {c: {t: {str(h): list(v) for h, v in holes.items()} for t, holes in tees.items()} for c, tees in course_holes_by_course.items()}
    json.dump(serializable, f)

print("done")
