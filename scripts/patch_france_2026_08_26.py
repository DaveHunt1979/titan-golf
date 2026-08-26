#!/usr/bin/env python3
"""
One-off patch to the France source workbook, applied 2026-08-26 after
Firecrawl research resolved two of the three courses needs_review.csv had
flagged (Belle Dune, Golf du Champ de Bataille). Golf de Saint-Saens stays
held out — no source has a hole-by-hole card matching its authoritative
rating (Par 70, 5832m, CR 72.3 / Slope 138, confirmed via the official club
site and the FFGolf-linked scoring calculator).

Belle Dune: the existing Blanc row's hole 15 (Par 6, 352yd, SI 1) is a real
signature hole, confirmed independently via TripAdvisor and the course's own
listing on golfencotedopale.com -- not a data error. It was only ever
excluded because merge_course_master.py capped par at 5 (now fixed to 6).
Left untouched here; adding Jaune as a second complete rated tee
(golfify.io, cross-checked against the FFGolf-linked calculator, which
independently confirms the same 71.6/125 rating).

Golf du Champ de Bataille: replaced entirely with the official club
scorecard (golf-champdebataille.com/carte-score, scanned image), which gives
all four tees (Blanc/Jaune/Bleu/Rouge) complete, including separate men's/
women's ratings for Jaune/Bleu/Rouge -- superseding the old Blanc-only,
16-of-18-hole 18birdies data.

Usage:
    python3 scripts/patch_france_2026_08_26.py
"""
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
FRANCE_PATH = ROOT / "screenshots" / "Golf courses" / "euro courses" / "France_GOLFBREAKS_YGT_FRANCE_FINAL_AUDIT_PROGRESS_59.xlsx"

COURSES_HDR = ["Course ID", "Official Course Name", "Club Name", "Region", "Town/Area",
               "Course / Layout", "Holes", "Verification Status", "Primary Source",
               "Last Verified", "Source Quality"]
TEES_HDR = ["Course ID", "Official Course Name", "Tee Name", "Gender", "Total Par",
            "Total Length (yds)", "Course Rating", "Slope Rating", "Front 9 Rating",
            "Back 9 Rating", "Source"]
HOLES_HDR = ["Course ID", "Official Course Name", "Tee Name", "Gender", "Hole",
             "Yardage", "Par", "Stroke Index", "Source"]

GOLFIFY_BELLE_DUNE = "https://www.golfify.io/courses/golf-de-belle-dune"
OFFICIAL_CHAMP_BATAILLE = "https://www.golf-champdebataille.com/carte-score"

# ---------------------------------------------------------------------------
# Belle Dune -- Jaune (new second tee; Blanc is untouched, already present)
# ---------------------------------------------------------------------------
BELLE_DUNE_JAUNE = {
    "course_id": "FRA_BELLE_DUNE_JAUNE",
    "name": "Belle Dune — Jaune",
    "club": "Golf de Belle Dune",
    "region": "North East France",
    "town": "Fort-Mahon-Plage",
    "layout": "Belle Dune",
    "par": 72,
    "total_distance": 5482,
    "course_rating": 71.6,
    "slope_rating": 125,
    "holes": [  # (hole, yardage, par, SI)
        (1, 328, 4, 12), (2, 323, 4, 7), (3, 126, 3, 18), (4, 419, 5, 16),
        (5, 349, 4, 9), (6, 145, 3, 14), (7, 352, 4, 5), (8, 366, 4, 3),
        (9, 279, 4, 11), (10, 493, 5, 4), (11, 350, 4, 6), (12, 131, 3, 13),
        (13, 441, 5, 2), (14, 127, 3, 17), (15, 330, 5, 1), (16, 232, 4, 15),
        (17, 323, 4, 8), (18, 368, 4, 10),
    ],
}

# ---------------------------------------------------------------------------
# Champ de Bataille -- full replacement, all 4 tees, official scanned card
# ---------------------------------------------------------------------------
CDB_PAR = [4, 3, 4, 5, 3, 4, 4, 5, 4, 5, 4, 3, 4, 4, 4, 5, 3, 4]
CDB_SI = [9, 16, 3, 1, 18, 17, 11, 4, 6, 12, 8, 14, 2, 13, 5, 10, 15, 7]

CHAMP_BATAILLE_TEES = [
    {
        "course_id": "FRA_CHAMP_BATAILLE_BLANC", "tee_name": "Blanc", "gender": "M",
        "name": "Golf du Champ de Bataille — Blanc",
        "total_distance": 5993, "course_rating": 72.3, "slope_rating": 144,
        "yardage": [350, 157, 317, 502, 152, 320, 349, 455, 374, 467, 337, 174, 349, 350, 398, 457, 136, 349],
    },
    {
        "course_id": "FRA_CHAMP_BATAILLE_JAUNE", "tee_name": "Jaune", "gender": "M",
        "name": "Golf du Champ de Bataille — Jaune",
        "total_distance": 5640, "course_rating": 70.2, "slope_rating": 141,
        "yardage": [332, 157, 285, 492, 134, 301, 304, 434, 358, 447, 304, 159, 343, 306, 374, 444, 131, 335],
    },
    {
        "course_id": "FRA_CHAMP_BATAILLE_JAUNE", "tee_name": "Jaune", "gender": "F",
        "name": "Golf du Champ de Bataille — Jaune",
        "total_distance": 5640, "course_rating": 76.0, "slope_rating": 146,
        "yardage": [332, 157, 285, 492, 134, 301, 304, 434, 358, 447, 304, 159, 343, 306, 374, 444, 131, 335],
    },
    {
        "course_id": "FRA_CHAMP_BATAILLE_BLEU", "tee_name": "Bleu", "gender": "M",
        "name": "Golf du Champ de Bataille — Bleu",
        "total_distance": 5103, "course_rating": 67.5, "slope_rating": 138,
        "yardage": [297, 142, 267, 482, 114, 277, 232, 417, 274, 427, 294, 136, 285, 292, 364, 423, 121, 259],
    },
    {
        "course_id": "FRA_CHAMP_BATAILLE_BLEU", "tee_name": "Bleu", "gender": "F",
        "name": "Golf du Champ de Bataille — Bleu",
        "total_distance": 5103, "course_rating": 72.8, "slope_rating": 144,
        "yardage": [297, 142, 267, 482, 114, 277, 232, 417, 274, 427, 294, 136, 285, 292, 364, 423, 121, 259],
    },
    {
        "course_id": "FRA_CHAMP_BATAILLE_ROUGE", "tee_name": "Rouge", "gender": "M",
        "name": "Golf du Champ de Bataille — Rouge",
        "total_distance": 4486, "course_rating": 64.5, "slope_rating": 129,
        "yardage": [261, 71, 189, 436, 102, 233, 232, 401, 267, 376, 285, 127, 248, 217, 304, 364, 121, 252],
    },
    {
        "course_id": "FRA_CHAMP_BATAILLE_ROUGE", "tee_name": "Rouge", "gender": "F",
        "name": "Golf du Champ de Bataille — Rouge",
        "total_distance": 4486, "course_rating": 69.4, "slope_rating": 130,
        "yardage": [261, 71, 189, 436, 102, 233, 232, 401, 267, 376, 285, 127, 248, 217, 304, 364, 121, 252],
    },
]


def assert_checksums():
    front, back = sum(CDB_PAR[:9]), sum(CDB_PAR[9:])
    assert front == 36 and back == 36 and front + back == 72, "Champ de Bataille par mismatch"
    for t in CHAMP_BATAILLE_TEES:
        yards = t["yardage"]
        assert len(yards) == 18, f"{t['tee_name']}/{t['gender']}: expected 18 holes"
        assert sum(yards) == t["total_distance"], f"{t['tee_name']}/{t['gender']}: yardage sum mismatch"
    jy = [h[1] for h in BELLE_DUNE_JAUNE["holes"]]
    jp = [h[2] for h in BELLE_DUNE_JAUNE["holes"]]
    assert sum(jy) == BELLE_DUNE_JAUNE["total_distance"], "Belle Dune Jaune yardage sum mismatch"
    assert sum(jp) == BELLE_DUNE_JAUNE["par"], "Belle Dune Jaune par sum mismatch"


def find_row(ws, header, course_id_col="Course ID"):
    idx = {h: i for i, h in enumerate(header)}
    return idx


def main():
    assert_checksums()

    wb = openpyxl.load_workbook(FRANCE_PATH)
    courses_ws = wb["Courses"]
    tees_ws = wb["Tee Ratings"]
    holes_ws = wb["Hole Data"]

    # --- Remove the old, incomplete Champ de Bataille rows (Blanc-only, 16 holes) ---
    def strip_course_id(ws, course_id, id_col_name="Course ID"):
        header = [c.value for c in ws[1]]
        col = header.index(id_col_name)
        rows_to_delete = [r for r in range(ws.max_row, 1, -1) if ws.cell(r, col + 1).value == course_id]
        for r in rows_to_delete:
            ws.delete_rows(r)
        return len(rows_to_delete)

    removed = (
        strip_course_id(courses_ws, "FRA_CHAMP_BATAILLE_WHITE")
        + strip_course_id(tees_ws, "FRA_CHAMP_BATAILLE_WHITE")
        + strip_course_id(holes_ws, "FRA_CHAMP_BATAILLE_WHITE")
    )
    print(f"Removed {removed} old FRA_CHAMP_BATAILLE_WHITE rows (incomplete, superseded)")

    # --- Belle Dune: add Jaune tee ---
    courses_ws.append([
        BELLE_DUNE_JAUNE["course_id"], BELLE_DUNE_JAUNE["name"], BELLE_DUNE_JAUNE["club"],
        BELLE_DUNE_JAUNE["region"], BELLE_DUNE_JAUNE["town"], BELLE_DUNE_JAUNE["layout"], 18,
        "Complete / full 18-hole scorecard + CR/Slope verified", GOLFIFY_BELLE_DUNE,
        "2026-08-26", "Published full scorecard + numeric CR/Slope",
    ])
    tees_ws.append([
        BELLE_DUNE_JAUNE["course_id"], BELLE_DUNE_JAUNE["name"], "Jaune", "M",
        BELLE_DUNE_JAUNE["par"], BELLE_DUNE_JAUNE["total_distance"],
        BELLE_DUNE_JAUNE["course_rating"], BELLE_DUNE_JAUNE["slope_rating"],
        None, None, GOLFIFY_BELLE_DUNE,
    ])
    for hole, yardage, par, si in BELLE_DUNE_JAUNE["holes"]:
        holes_ws.append([
            BELLE_DUNE_JAUNE["course_id"], BELLE_DUNE_JAUNE["name"], "Jaune", "M",
            hole, yardage, par, si, GOLFIFY_BELLE_DUNE,
        ])

    # --- Champ de Bataille: full 4-tee replacement ---
    seen_course_ids = set()
    for t in CHAMP_BATAILLE_TEES:
        if t["course_id"] not in seen_course_ids:
            seen_course_ids.add(t["course_id"])
            courses_ws.append([
                t["course_id"], t["name"], "Golf du Champ de Bataille", "Normandy",
                "Le Neubourg", "Champ de Bataille", 18, "COMPLETE", OFFICIAL_CHAMP_BATAILLE,
                "2026-08-26", "Official club scorecard, all 4 tees, men's/women's ratings",
            ])
        tees_ws.append([
            t["course_id"], t["name"], t["tee_name"], t["gender"], 72,
            t["total_distance"], t["course_rating"], t["slope_rating"],
            None, None, OFFICIAL_CHAMP_BATAILLE,
        ])
        for i in range(18):
            holes_ws.append([
                t["course_id"], t["name"], t["tee_name"], t["gender"],
                i + 1, t["yardage"][i], CDB_PAR[i], CDB_SI[i], OFFICIAL_CHAMP_BATAILLE,
            ])

    wb.save(FRANCE_PATH)
    print(f"Saved {FRANCE_PATH}")
    print("Added: Belle Dune Jaune (1 tee, 18 holes)")
    print("Added: Champ de Bataille Blanc/Jaune(M+F)/Bleu(M+F)/Rouge(M+F) (7 tee rows, 126 hole rows)")


if __name__ == "__main__":
    main()
